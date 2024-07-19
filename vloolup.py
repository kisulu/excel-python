import openpyxl


wb = openpyxl.load_workbook("cities_changes.xlsx")
wb1 = openpyxl.load_workbook("samplewb.xlsx")

sheet_1 = wb["Sheet1"]
sheet_2 = wb1["sample_sheet"]


for i in sheet_2.iter_rows():
    id = i[0].value
    row_number = i[0].row
    # print(f'{row_number}  : {id}')
    for j in sheet_1.iter_rows():
        if j[0].value == id:
##            print (j[1].value)
##            print (j[2].value)
##            print (j[3].value)
            sheet_2.cell(row=row_number, column=7).value =j[1].value
            sheet_2.cell(row=row_number, column=8).value =j[2].value
            sheet_2.cell(row=row_number, column=9).value =j[3].value
        else:
            print("terminate")

sheet_2.save("newcities.xlsx")
