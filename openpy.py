import openpyxl
from openpyxl.workbook import workbook
from openpyxl.styles import PatternFill

wk = openpyxl.load_workbook("cities.xlsx")
wk2 = openpyxl.load_workbook("cities1.xlsx")

fill_style=PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid")

sheet = wk['Sheet1']
sheet2 = wk2['Sheet1']

for row in sheet.iter_rows():
    for cell in row:
        current_cell_value = cell.value # print(cell.value)
        cell_location = cell.coordinate # print(cell.cordinate)

        if current_cell_value != sheet2[cell_location].value:
            # print(sheet2[cell_location])
            cell.fill = fill_style

wk.save("cities_changes.xlsx")


