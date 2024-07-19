import openpyxl

from openpyxl.workbook import workbook
from openpyxl.styles import Font, Color

# Background color of cell/s
from openpyxl.styles import PatternFill

# Format 
from openpyxl.styles import numbers

wb = openpyxl.load_workbook("samplewb.xlsx")
wb1 = openpyxl.load_workbook("balance.xlsx")
ws = wb['sample_sheet']
ws1= wb1["Sheet1"]

_style = Font(color="1A4FDF", name="Cooper Black", size=14, italic=True, bold=True) # property
a1 = ws["A1"]
a1.font = _style

font_style = Font(color="DB3B22", name="Reem Kufi", size=12, italic=True, bold=True, 
                 underline="single", strikethrough=False ) # property
for i in range(2,61):
	ws.cell(row=i,column=3).font = font_style

fill_pattern = PatternFill(patternType='solid', fgColor="C64747") #variable /object to store
ws["B5"].fill = fill_pattern


# Automation of formulas
ws["A62"] = "SUM"
ws['C62'].value = "=SUM(C2:C60)"
ws["C62"].font = Font(bold=True)
ws["A62"].font = Font(bold=True)

ws["A63"] = "AVERAGE"
ws["C63"] = "=AVERAGE(C2:C60)" 
ws["C63"].font = Font(bold=True)
ws["A63"].font = Font(bold=True)


# USER DEFINED FORMULAS
ws['M1'].value = "Balance after a year"
ws["M1"].font = Font(name="Calibri", bold=True, size=11, color="000000")
# ws["M1"].fill = PatternFill(patternType='solid', fgColor="Blue")

ws["M1"].number_format = numbers.FORMAT_TEXT

# iterate a row
for i in range(1, 12):
	print(ws.cell(row=1,column=i).value)

# iterate a column
for i in range(2, 10):
	balance = ws1.cell(row=i,column=2).value
	interest = ws1.cell(row=i,column=3).value
	final_balance = (balance*interest)+balance

	ws.cell(row=i,column=13).value = final_balance
	
	# print(ws.cell(row=i,column=4).value)

wb.save("font_cities.xlsx")
