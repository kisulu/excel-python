import openpyxl
from openpyxl.style import Font

wb0 = openpyxl.load_workbook("file_0")
wb1 = openpyxl.load_workbook("file_1")

ws0 = wb0[]
ws1 = wb1[]

# row count for first wb
is_data = True
row_count = 1

while is_data:
    row_count += 1
    data = ws0.cell(row=row_count,column=1).value
    if data ==None:
        is_data = False

print(row_count)

# row count for second wb
while is_data:
    row_count += 1
    data = ws1.cell(row=row_count,column=1).value
    if data ==None:
        is_data = False

print(row_count)



for i in ramge(1, row_count):
    row_data = {}
    row_data["id"] = ws0.cell(row=i, column=1).value
    row_data['todays_purchase'] =ws0.cell(row=i, column=2).value
    row_data["todays_reward"] = ws0.cell(row=i, column=3)
    todays_data.append(row_data)

# print(todays_data)

for i in range(2, wb1):
    id=ws1.cell(row=i, column=1).value
    for row in todays_data:
        if row['id'] ==id:
            todays_purchase =int(row["todays_purchase"])
            todays_reward=row['todays_reward']

            # get from master sheet
            total_purchase = ws0.cell(row=i, column=6).value
            total_reward = ws0.cell(row=i,column=7).value

            # new wb into total data
            new_total_purchase = total_purchase + todays_purchase
            new_total_reward = total_reward + todays_reward

            ws0.cell(row=i, column=6).value = new_total_purchase
            ws0.cell(row=i, column=7).value = new_total_reward

wb0.save("new_execel_file")
        
# create a blank workbook object
daily_report = openpyxl.Workbook()
ws= daily_report.active

# get header
is_data =True
column_count=1
header-values=[]

while True:
    column_count += 1
    data = ws0.cell(row=1,column=column_count)
    if data != None:
        header_value.append(data)
    else:
        is_data =False
        
header_style = Font(name="Times New Roman", size=12, bold=True)
for i, col_name in enumarate(header_values):
    col_index=i+1
    ws.cell(row=1, column=col_index).value = col_name
    ws.cell(row=1,column=col_index).font = header_style

IDs =[]
for data in todays_data:
    IDs.append(data['id'])

IDs.pop(0)
print(IDs)

final_data = []
for i in range(2, row_count):
    id = ws0.cell(row=i, column=1).value
    if id in IDs:
        lst = []
        for j in range(2, 8):
            lst.append(ws0.cell(row=1, coulumn=j).value)
        final_data.append[lst]

print(final_data)
        
for data in final_data:
    ws.append(data)
    
daily_report.save("report.xlsx")

