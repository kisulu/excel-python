import openpyxl

wb = openpyxl.load_workbook("samplewb.xlsx")
ws = wb.active

is_wb = True
row_count = 1

while is_wb:
    row_count =+ 1
    latitude = ws.cell(row=row_count, column=9).value
    if latitude != None:
        # print(latitude.strip('.'))
        ws.cell(row=row_count, column=2).value = latitude.strip('.')
    else:
        is_wb = False


wb.save("sample.xlsx")
