import openpyxl
from openpyxl.workbook import workbook
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook("cities.xlsx")


ws = wb.active

column_a = ws['2']
print(column_a)

tup0 = ("<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>, <Cell 'Sheet1'.D1>")
tup = f"({tup0})"
print(tup)



for cell in column_a:
    print(f'{cell.value}', end=":") #f'{cell.value},{cell.coordinate}'



